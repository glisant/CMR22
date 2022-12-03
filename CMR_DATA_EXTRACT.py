from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from natsort import natsorted
import re

import copy
import pandas as pd
pd.set_option('display.max_rows', None)
import statistics
in_empty=True
BIOMARKERS_INPUT=['Biomarker: End-Systole Mean Intensity','Biomarker: End-Systole Intensity - Phase Normalized',
              'Biomarker: End-Systole Intensity - Series Normalized','Biomarker: End-Systole O2 Relative to Lv Bloodpool',
              'Biomarker: End-Systole O2 Relative to Rv Bloodpool','Biomarker: End-Systole O2 Relative to Lv & Rv Bloodpools'
              ,'Biomarker: End-Diastole Mean Intensity','Biomarker: End-Diastole Intensity - Phase Normalized'
              ,'Biomarker: End-Diastole Intensity - Series Normalized','Biomarker: End-Diastole O2 Relative to Lv Bloodpool'
              ,'Biomarker: End-Diastole O2 Relative to Rv Bloodpool','Biomarker: End-Diastole O2 Relative to Lv & Rv Bloodpools']

input_files=['GLOBAL.xlsx','ENDO.xlsx','EPI.xlsx']
phase_file='phases.xlsx'

def get_phase(bio_in,phase_file,i_files,required=[1,1,1]):
  
  wb_input=load_workbook(phase_file,read_only=True)
  
  all_patient_wb=load_workbook(i_files[0],read_only=True).sheetnames
  phase_info={}

  possible_reqs=['Baseline', '0s', '30s']
  reqs=[]
  for _, req in enumerate(required):
    if req:
      reqs.append((possible_reqs[_],[]))
  reqs=tuple(reqs)
  reqs_f = {k: v for k, v in reqs}
  #deep_reqs=copy.deep(reqs_f)
  
  for _,patient_phase in enumerate(wb_input.sheetnames):
      
      try:
        patient=all_patient_wb[_]
      except:
        continue

      if patient==None:
        continue
      #except:
        #continue
      ws_input=wb_input[patient_phase]
      phase_info[patient]={}
      
      phase_info[patient]['ES']=copy.deepcopy(reqs_f)
      phase_info[patient]['ED']=copy.deepcopy(reqs_f)
      

      for rowid, row in enumerate(ws_input.iter_rows(min_row=0, max_col=7, max_row=4)):       

        if rowid==0:
          phase_names=row
        if rowid==1:
            pat_rows=row
        
        if rowid==2 or rowid==3:
            for _,cell in enumerate(row):
                phase_id=cell.value
                if '`' in str(cell.value):
                    phase_id=phase_id.replace('`','')

                if phase_id==None or phase_id=='':
                  phase_id='MISSING'
                if pat_rows[_].value=='ES':
                    curr_phase=phase_names[_].value
                    
                    if curr_phase==None:
                      curr_phase=phase_names[_-1].value
                    #print(curr_phase,'ES',phase_id,patient)
                    
                    phase_info[patient]['ES'][curr_phase].append(phase_id)
                    #print(phase_info[patient]['ES'][curr_phase])
                    #o=input()
                    #print(patient,biomarker,curr_phase,phase_id)
                    #o=input()

  
                elif pat_rows[_].value=='ED':
                    curr_phase=phase_names[_].value
                    if curr_phase==None:
                      curr_phase=phase_names[_-1].value
                    
                    phase_info[patient]['ED'][curr_phase].append(phase_id)
                    #print(phase_info[patient]['ED'][curr_phase])
                    #o=input()
                    #print(patient,biomarker,curr_phase,phase_id)
                    #o=input()
      #print(phase_info)
      #o=input('Enter to Continue')
  #print(phase_info[patient]['ES']['0s'])
  #print(phase_info[patient]['ED']['0s'])
  #print(phase_info[patient]['ES']['0s']==phase_info[patient]['ED']['0s'])
 # o=input()
  #print(phase_info)
  #1/0
  return phase_info

def get_raw_all(ifile,phase_info,rq_phase):
  input_files=ifile
  all_ex={}
  for filen in input_files:
      name=filen.replace('.xlsx', "")
      print('Currently working with:',name)
      #lab0.config(text=f'Currently working with: {name}...')
      all_ex[name]=get_global(filen,phase_info,rq_phase)
  return(all_ex)

def get_global(input_file,phase_info,r_phase):
    global BIOMARKERS_INPUT
    global_ex={}
    wb_input=load_workbook(input_file,read_only=True)
    for sheet in phase_info:
        ws_input=wb_input[sheet]
        global_ex[sheet]=get_phases_AHA_Segments(BIOMARKERS_INPUT,ws_input,sheet=input_files.index(input_file),ph_info=phase_info,patient=sheet,rq_phase=r_phase)
    return global_ex

def get_phases_AHA_Segments(pi,wsi,sheet=0,ph_info=[],patient='',rq_phase=[1,1,1]):
    coords=ph_info
    phases_input=pi
    ws_input=wsi
    start=0
    if sheet==1:
        start=90
    elif sheet==2:
        start=180
        
    worksheet_dict={'Norm':{'0s':{},'30s':{}},'Base':{}}
    bi=0
    next_u='start'

    for rowid, row in enumerate(ws_input.iter_rows(min_row=0, max_col=90, max_row=197)):

        if row[0].value!=None and row[0].value[0:9]=='Biomarker':
            #For Debugging purposes:
            #print(f'Working with: {phases_input[bi]}')

            working_marker_ES=phases_input[bi]
            
            working_marker_ED=phases_input[bi+6]
            
            bi+=1
            next_u=rowid+5
            worksheet_dict['Norm']['0s'][working_marker_ES]=[]
            worksheet_dict['Norm']['30s'][working_marker_ES]=[]
            worksheet_dict['Norm']['0s'][working_marker_ED]=[]
            worksheet_dict['Norm']['30s'][working_marker_ED]=[]
            #worksheet_dict['Norm'][working_marker_ES]=[]
            #worksheet_dict['Norm'][working_marker_ED]=[]
        if rowid==next_u:
            next_u+=1
            seg=row[0].value
            
            seg=int(re.search(r'\d+', seg).group())
            #seg=seg.strip()
            #seg = seg.replace('Segment', "")
            #seg = seg.replace(' ', "")
            #seg = seg.replace('AHA', "")
            if seg>6:
              segment_id=1
            elif seg<7:
              segment_id=0
            if seg==6:
              next_u=rowid+5
            if seg==12:
              next_u='start'
            
            col_for_ex=seg
            #print(patient,working_marker_ES,segment_id, coords[patient][working_marker_ES])
            #print(coords)
            #print(patient)
            #print(coords[patient]['ES']['0s'])
            #print(coords[patient]['ED']['0s'])
            #print(coords[patient]['ES']['Baseline']==coords[patient]['ED']['Baseline'])
            #o=input()
            key_here='MISSING'
            if '0s' in coords[patient]['ES'] and len(coords[patient]['ES']['0s'])>1:
              key_here=coords[patient]['ES']['0s'][segment_id]
           
            #print(0,patient,working_marker_ES,segment_id,key_here)
            #o=input()
            #print(key_here)
            if key_here!='MISSING':
              _0s_ES=row[key_here].value
            else:
              _0s_ES='MISSING'
            
            if '30s' in coords[patient]['ES'] and len(coords[patient]['ES']['30s'])>1:
              key_here=coords[patient]['ES']['30s'][segment_id]
            #print(30,patient,working_marker_ES,segment_id,key_here)
            #o=input()
            if key_here!='MISSING':
              _30s_ES=row[key_here].value
            else:
              _30s_ES='MISSING'
            if '0s' in coords[patient]['ED'] and len(coords[patient]['ED']['0s'])>1:
              key_here=coords[patient]['ED']['0s'][segment_id]
            #print(100,patient,working_marker_ED,segment_id,key_here)
           # o=input()
            if key_here!='MISSING':
              _0s_ED=row[key_here].value
            else:
              _0s_ED='MISSING'
            if '30s' in coords[patient]['ED'] and len(coords[patient]['ED']['30s'])>1:
              key_here=coords[patient]["ED"]['30s'][segment_id]
            #print(3030,patient,working_marker_ED,segment_id,key_here)
            #o=input()
            if key_here!='MISSING':
              _30s_ED=row[key_here].value
            else:
              _30s_ED='MISSING'

            if _0s_ES==None:
                _0s_ES='MISSING'
            if _30s_ES==None:
                _30s_ES='MISSING'
            if _0s_ED==None:
                _0s_ED='MISSING'
            if _30s_ED==None:
                _30s_ED='MISSING'

            #print(f'Keys found: {_0s_ES} {_30s_ES} {_0s_ED} {_30s_ED}')
                
            #worksheet_dict['Norm'][working_marker_ES].append([_0s_ES,0])
            #worksheet_dict['Norm'][working_marker_ES].append([_30s_ES,30])
            
            #worksheet_dict['Norm'][working_marker_ED].append([_0s_ED,0])
            #worksheet_dict['Norm'][working_marker_ED].append([_30s_ED,30])

            worksheet_dict['Norm']['0s'][working_marker_ES].append(_0s_ES)
            #print(working_marker_ES,_0s_ES)
            worksheet_dict['Norm']['30s'][working_marker_ES].append(_30s_ES)
            
            worksheet_dict['Norm']['0s'][working_marker_ED].append(_0s_ED)
            #print(working_marker_ED,_0s_ED)
            #o=input()
            worksheet_dict['Norm']['30s'][working_marker_ED].append(_30s_ED)

    bi=0
    next_u='start'
    
    rows_to_check=[578,950]
    if rq_phase[1]==0:
      rows_to_check=[0,197]

    for rowid, row in enumerate(ws_input.iter_rows(min_row=rows_to_check[0], max_col=90, max_row=rows_to_check[1])):
        
        if row[0].value!=None and row[0].value[0:9]=='Biomarker':
            if bi==6:
              break
            
            #print(f'Working with: {phases_input[bi]}')
            working_marker_ES=phases_input[bi]
            try:
              working_marker_ED=phases_input[bi+6]
            except:
              working_marker_ED=phases_input[-1]
            #print(bi)
            bi+=1
            next_u=rowid+5

            worksheet_dict['Base'][working_marker_ES]=[]
            worksheet_dict['Base'][working_marker_ED]=[]
        if rowid==next_u:
            next_u+=1
            seg=row[0].value
            
            seg=int(re.search(r'\d+', seg).group())
           
            if seg>6:
              segment_id=1
            elif seg<7:
              segment_id=0
            if seg==6:
              next_u=rowid+5
            if seg==12:
              next_u='start'
            

            col_for_ex=seg
            key_here='MISSING'
            if 'Baseline' in coords[patient]['ES'] and len(coords[patient]['ES']['Baseline'])>1:
              key_here=coords[patient]['ES']['Baseline'][segment_id]
            

            if key_here!='MISSING':
              Base_ES=row[key_here].value
            else:
              Base_ES='MISSING'
            if 'Baseline' in coords[patient]['ED'] and len(coords[patient]['ED']['Baseline'])>1:
              key_here=coords[patient]['ED']['Baseline'][segment_id]
            if key_here!='MISSING':
              Base_ED=row[key_here].value
            else:
              Base_ED='MISSING'
            
            if Base_ES==None:
                Base_ES='MISSING'
            if Base_ED==None:
                Base_ED='MISSING'     

            #print(f'Keys found Baseline: ES: {Base_ES} ED:{Base_ED} Patient: {patient} BIO:{working_marker_ES}')   
            #wuu=input('Press Enter to Continue')    

            worksheet_dict['Base'][working_marker_ES].append(Base_ES)
           
            worksheet_dict['Base'][working_marker_ED].append(Base_ED)
    return worksheet_dict

def average(user_in):
  try:
    return sum(user_in)/len(user_in)
  except:
    return 'MISSING'

def average_Basal_slice(segments):
  """
  Returns average of basal slices
  """
  id_f='Basal Slice'
  return average(segments[0:6])

def average_Mid_slice(segments):
  """
  Returns average of mid slices
  """
  id_f='Mid Slice'
  return average(segments[6:])

def average_Global(segments):
  """
  Returns average of basal+mid slices
  """
  id_f='Global'
  return average(segments)

def average_LAD(segments):
  """
  Returns average of LAD segments
  """
  id_f='LAD'
  keep_seg = [0,1,6,7]
  segments = [i for j, i in enumerate(segments) if j in keep_seg]
  return average(segments)

def average_RCA(segments):
  """
  Returns average of RCA segments
  """
  id_f='RCA'

  keep_seg = [2,3,8,9]
  segments = [i for j, i in enumerate(segments) if j in keep_seg]
  return average(segments)

def average_Circumflex(segments):
  """
  Returns average of LCX (circumflex) segments
  """
  id_f='LCX'
  keep_seg = [4,5,10,11]
  segments = [i for j, i in enumerate(segments) if j in keep_seg]
  return average(segments)

def perc_change_OS(segments1, segments2):
  output=[]
  for i in range(12):
    try:
      output.append(((segments2[i]-segments1[i])/segments1[i])*100)
    except:
      output.append('MISSING')
  return output

def MORE_SD(segments1,segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns STD
  """
  id_f='MORE SD'
  try :
    return(round(statistics.stdev(segments1+segments2),4))
  except:
    return('MISSING')

def MORE_Range(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and 
  returns range
  """
  id_f='MORE RANGE'
  try :
    return(max(segments1+segments2)-min(segments1+segments2))
  except:
    return('MISSING')

def trans_reg_het(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the 
  sum of differences
  """

  id_f='Transmural Regional Heterogeneity'
  differences=[]
  for i in range(12):
    try:
      differences.append(segments2[i]-segments1[i])
    except:
      return 'MISSING'
  return (sum(differences))

def radial_het(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the absolute 
  sum of differences
  """
  id_f='Radial Heterogeneity'
  differences=[]
  for i in range(12):
    try:
      differences.append(abs(segments2[i]-segments1[i]))
    except:
      return 'MISSING'
  return (sum(differences))

def circumferential_het(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the pair wise
  difference
  """
  id_f='Circumferential Heterogeneity'
  try:
    differences=[segments1[0]-segments1[-1],segments2[0]-segments2[-1]]
  except:
    return 'MISSING'
  for i in range(11):
    try:
      differences.append((segments1[i]-segments1[i+1]))
      differences.append((segments2[i]-segments2[i+1]))
    except:
      return 'MISSING'
  return (sum(differences))

def circumferential_het_abs(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the pair wise
  absolute difference
  """
  id_f='ABS Circumferential Heterogeneity'
  try:
    differences=[abs(segments1[0]-segments1[-1]),abs(segments2[0]-segments2[-1])]
  except TypeError:
    return 'MISSING'
  for i in range(11):
    try:
      differences.append(abs(segments1[i]-segments1[i+1]))
      differences.append(abs(segments2[i]-segments2[i+1]))
    except:
      return 'MISSING'
  return (sum(differences))

def trans_long_het(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the 
  pair wise sum of differences of base-slice 
  """
  id_f='Transmural Longitudinal Heterogeneity'
  differences=[]
  for i in range(6):
    try:
      differences.append(segments2[i]-segments2[i+6])
      differences.append(segments1[i]-segments1[i+6])
    except:
      return 'MISSING'
  return (sum(differences))

def long_het(segments1, segments2):
  """
  Takes Epi and Endo MORE Segmental% change OS values and returns the 
  pair wise sum of the absolute differences of base-slice 
  """
  id_f=id_f='Longitudinal Heterogeneity'
  differences=[]
  for i in range(6):
    try:
      differences.append(abs(segments2[i]-segments2[i+6]))
      differences.append(abs(segments1[i]-segments1[i+6]))
    except:
      return 'MISSING'
  return (sum(differences))

def get_all_regional(segments1,segments2):
  
  dic_export=[('MORE SD',MORE_SD(segments1,segments2)),
  ('MORE Range',MORE_Range(segments1, segments2)),
  ('Transmural Regional Heterogeneity',trans_reg_het(segments1, segments2)),
  ('Radial Heterogeneity',radial_het(segments1, segments2)),
  ('Circumferential Heterogeneity',circumferential_het(segments1, segments2)),
  ('ABS Circumferential Heterogeneity',circumferential_het_abs(segments1, segments2)),
  ('Transmural Longitudinal Heterogeneity',trans_long_het(segments1, segments2)),
  ('Longitudinal Heterogeneity',long_het(segments1, segments2))]

  dic_export=tuple(dic_export)
  dic_export = {k: v for k, v in dic_export}
  return dic_export
  
def get_basic(segments):

  dic_export=[('Basal Slices',average_Basal_slice(segments)),
  ('Mid Slices',average_Mid_slice(segments)),
  ('MORE Slices',average_Global(segments))]

  dic_export=tuple(dic_export)
  dic_export = {k: v for k, v in dic_export}
  return dic_export

def get_basic_coro(segments):

  dic_export=[('Basal Slices',average_Basal_slice(segments)),
  ('Mid Slices',average_Mid_slice(segments)),
  ('MORE Slices',average_Global(segments)),('LAD',average_LAD(segments)),
  ('RCA',average_RCA(segments)),
  ('LCX',average_Circumflex(segments))]

  dic_export=tuple(dic_export)
  dic_export = {k: v for k, v in dic_export}
  return dic_export
def get_coro(segments):

  dic_export=[
  ('LAD',average_LAD(segments)),('RCA',average_RCA(segments)),
  ('LCX',average_Circumflex(segments))]

  dic_export=tuple(dic_export)
  dic_export = {k: v for k, v in dic_export}
  return dic_export
  
def ready_export_to_excel(all_ex):
  """
  Grabs a dictionary of AHA segments separated as follows:
  DIC[FILE_TYPE][PATIENT][BREATHHOLD][BASE/NORM][BIOMARKER]=[12 AHA SEGMENTS]

  Ex:
  all_ex['GLOBAL']['HDP 020 LUGLA']['Norm']['0s']['Biomarker: End-Systole Mean Intensity']
  
  """
  final_export={}
  different_files=list(all_ex.keys())

  patients=list(all_ex[different_files[0]].keys())
  breath_holds=list(all_ex[different_files[0]][patients[0]].keys())
  b_n_types=list(all_ex[different_files[0]][patients[0]][breath_holds[0]].keys())
  markers=list(all_ex[different_files[0]][patients[0]][breath_holds[0]][b_n_types[0]].keys())

  for patient in patients:
    final_export[patient]={}
    for marker in markers:
      final_export[patient][marker]={}
      for file in different_files:
        final_export[patient][marker][file]={}
        
        if 'Base' in all_ex[file][patient] and len(list(all_ex[file][patient]['Base'].keys()))>0:
          segments_Base=all_ex[file][patient]['Base'][marker]  
          final_export[patient][marker][file]['Base Segments']=segments_Base
          final_export[patient][marker][file]['Base Ave']=get_basic(segments_Base)
          final_export[patient][marker][file]['Base Coro']=get_coro(segments_Base)
        
        if 'Norm' in all_ex[file][patient]:
          if '0s' in all_ex[file][patient]['Norm'] and len(list(all_ex[file][patient]['Norm']['0s'].keys()))>0:
            segments_0s=all_ex[file][patient]['Norm']['0s'][marker]
            final_export[patient][marker][file]['Norm 0s Segments']=segments_0s
            final_export[patient][marker][file]['Norm 0s Ave']=get_basic(segments_0s)
            final_export[patient][marker][file]['Norm 0s Coro']=get_coro(segments_0s)

          if '30s' in all_ex[file][patient]['Norm'] and len(list(all_ex[file][patient]['Norm']['30s'].keys()))>0:
            segments_30s=all_ex[file][patient]['Norm']['30s'][marker]
            final_export[patient][marker][file]['Norm 30s Segments']=segments_30s
            final_export[patient][marker][file]['Norm 30s Ave']=get_basic(segments_30s)
            final_export[patient][marker][file]['Norm 30s Coro']=get_coro(segments_30s)
          
        if '0s' in all_ex[file][patient]['Norm'] and '30s' in all_ex[file][patient]['Norm'] and len(list(all_ex[file][patient]['Norm']['30s'].keys()))>0:
          final_export[patient][marker][file]['MORE Segmental changeOS']=perc_change_OS(final_export[patient][marker][file]['Norm 0s Segments'], final_export[patient][marker][file]['Norm 30s Segments'])
          final_export[patient][marker][file]['MORE changeOS']=get_basic(final_export[patient][marker][file]['MORE Segmental changeOS'])
          final_export[patient][marker][file]['MORE Coro changeOS']=get_coro(final_export[patient][marker][file]['MORE Segmental changeOS'])
        
        if 'Base' in all_ex[file][patient] and len(list(all_ex[file][patient]['Base'].keys()))>0 and 'Norm' in all_ex[file][patient] :
          final_export[patient][marker][file]['HV MORE Segmental changeOS']=perc_change_OS(final_export[patient][marker][file]['Base Segments'], final_export[patient][marker][file]['Norm 0s Segments'])
          final_export[patient][marker][file]['HV MORE changeOS']=get_basic(final_export[patient][marker][file]['HV MORE Segmental changeOS'])
          final_export[patient][marker][file]['HV MORE Coro changeOS']=get_coro(final_export[patient][marker][file]['MORE Segmental changeOS'])
      
        if file==different_files[-1] and 'ENDO' in different_files and 'EPI' in different_files and len(list(all_ex[file][patient]['Base'].keys()))>0:
          final_export[patient][marker][file]['Regional Heterogeneity']=get_all_regional(final_export[patient][marker]['EPI']['MORE Segmental changeOS'],
          final_export[patient][marker]['ENDO']['MORE Segmental changeOS'])
  return final_export
def get_data_frame(all_ex,req_markers=[]):
  """
  Grabs a dictionary of AHA segments separated as follows:
  DIC[FILE_TYPE][PATIENT][BREATHHOLD][BASE/NORM][BIOMARKER]=[12 AHA SEGMENTS]

  Ex:
  all_ex['GLOBAL']['HDP 020 LUGLA']['Norm']['0s']['Biomarker: End-Systole Mean Intensity']
  
  """
  Regional={}
  final_export={'Patient':[],'Biomarker':[],'File':[],'Segment Type':[],'Segment Subtype':[],'Value':[]}
  different_files=list(all_ex.keys())
  patients=[]
  for file in different_files:
    for patient in all_ex[file]:
      if patient not in patients:
        patients.append(patient)
  segment_types=['HV MORE changeOS', 'MORE changeOS', 'Regional Heterogeneity']
  segment_sub_types=['Basal Slices','Mid Slices','MORE Slices','LAD','RCA','LCX']

  for patient in patients:
    Regional[patient]={}
    for file in different_files:
      Regional[patient][file]={}

      for s_type in all_ex[file][patient]:
        if s_type=="Norm": 
          for ss_type in all_ex[file][patient][s_type]:
            if ss_type=='0s':
              segment_type='0 Seconds'
              for marker in all_ex[file][patient][s_type][ss_type]:
                basic_averages=get_basic_coro(all_ex[file][patient][s_type][ss_type][marker])
                #print(all_ex[file][patient][s_type][ss_type][marker],marker)
                #o=input()
                for segment_sub_type in basic_averages:
                  final_export['Patient'].append(patient)
                  final_export['Biomarker'].append(marker)
                  final_export['File'].append(file)
                  final_export['Segment Type'].append(segment_type)
                  final_export['Segment Subtype'].append(segment_sub_type)
                  final_export['Value'].append(basic_averages[segment_sub_type])

                for valueid,value in enumerate(all_ex[file][patient][s_type][ss_type][marker]):
                  final_export['Patient'].append(patient)
                  final_export['Biomarker'].append(marker)
                  final_export['File'].append(file)
                  final_export['Segment Type'].append(segment_type)
                  final_export['Segment Subtype'].append(f'AHA {valueid+1}')
                  final_export['Value'].append(value)
            elif ss_type=='30s':
              segment_type='30 Seconds'
              for marker in all_ex[file][patient][s_type][ss_type]:
                basic_averages=get_basic_coro(all_ex[file][patient][s_type][ss_type][marker])
                for segment_sub_type in basic_averages:
                  final_export['Patient'].append(patient)
                  final_export['Biomarker'].append(marker)
                  final_export['File'].append(file)
                  final_export['Segment Type'].append(segment_type)
                  final_export['Segment Subtype'].append(segment_sub_type)
                  final_export['Value'].append(basic_averages[segment_sub_type])
                for valueid,value in enumerate(all_ex[file][patient][s_type][ss_type][marker]):
                  final_export['Patient'].append(patient)
                  final_export['Biomarker'].append(marker)
                  final_export['File'].append(file)
                  final_export['Segment Type'].append(segment_type)
                  final_export['Segment Subtype'].append(f'AHA {valueid+1}')
                  final_export['Value'].append(value)
        else:
          segment_type='Baseline'
          for marker in all_ex[file][patient][s_type]:
            basic_averages=get_basic_coro(all_ex[file][patient][s_type][marker])
            for segment_sub_type in basic_averages:
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append(segment_type)
              final_export['Segment Subtype'].append(segment_sub_type)
              final_export['Value'].append(basic_averages[segment_sub_type])
            
            for valueid,value in enumerate(all_ex[file][patient][s_type][marker]):
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append(segment_type)
              final_export['Segment Subtype'].append(f'AHA {valueid+1}')
              final_export['Value'].append(value)

          for marker in all_ex[file][patient][s_type]:
            HV_MCHG=perc_change_OS(all_ex[file][patient]['Base'][marker],all_ex[file][patient]['Norm']['0s'][marker])
            basic_averages=get_basic_coro(HV_MCHG)
            for valueid,value in enumerate(HV_MCHG):
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append('HV MORE changeOS')
              final_export['Segment Subtype'].append(f'AHA {valueid+1}')
              final_export['Value'].append(value)
            for segment_sub_type in basic_averages:
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append('HV MORE changeOS')
              final_export['Segment Subtype'].append(segment_sub_type)
              final_export['Value'].append(basic_averages[segment_sub_type])

            MCHOS=perc_change_OS(all_ex[file][patient]['Norm']['0s'][marker],all_ex[file][patient]['Norm']['30s'][marker])
            if file=='ENDO' or file=='EPI':
              Regional[patient][file][marker]=MCHOS
              
            basic_averages=get_basic_coro(MCHOS)
            for valueid,value in enumerate(MCHOS):
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append('MORE changeOS')
              final_export['Segment Subtype'].append(f'AHA {valueid+1}')
              final_export['Value'].append(value)
            for segment_sub_type in basic_averages:
              final_export['Patient'].append(patient)
              final_export['Biomarker'].append(marker)
              final_export['File'].append(file)
              final_export['Segment Type'].append('MORE changeOS')
              final_export['Segment Subtype'].append(segment_sub_type)
              final_export['Value'].append(basic_averages[segment_sub_type])
  
  for patient in patients:
    if 'ENDO' in all_ex:
      for marker in all_ex['ENDO'][patient]['Base']:
        clean_pat_endo=Regional[patient]['ENDO'][marker]
        clean_pat_epi=Regional[patient]['EPI'][marker]
        reg_heg=get_all_regional(clean_pat_epi,clean_pat_endo)
        for regh in reg_heg:
          final_export['Patient'].append(patient)
          final_export['Biomarker'].append(marker)
          final_export['File'].append('EPI+ENDO')
          final_export['Segment Type'].append('Regional Heterogeneity')
          final_export['Segment Subtype'].append(regh)
          final_export['Value'].append(reg_heg[regh])
  data_frame=pd.DataFrame.from_dict(final_export)
  return data_frame

def export_to_excel(f_export,filename='CMR_DATA_EXPORT',req_markers=[]):
  """
  Grabs a dictionary of AHA segments separated as follows:
  DIC[PATIENT][BIOMARKER][FILE][SEGMENT]=Values

  Ex:
  f_export['HDP 020 LUGLA']['Biomarker: End-Systole Mean Intensity']['GLOBAL']
  
  """
  rows_ex=[['Biomarker'],['Demographics'],['PatientID/OSCMR Value']]
  patients=list(f_export.keys())
  biomarkers=[]
  for patient in patients:
    for marker in f_export[patient]:
      if marker not in biomarkers:
        biomarkers.append(marker)
  #for _,marker in enumerate(biomarkers):
  #marker=marker.replace('Biomarker: ','')
  wb_output_1=Workbook(write_only=True)
  ws_out_1=wb_output_1.create_sheet(filename)
  files=['GLOBAL','ENDO','EPI']
  first_patient=True
  value_ids=['Base Segments', 'Base Ave', 'Base Coro', 'Norm 0s Segments',
            'Norm 0s Ave', 'Norm 0s Coro', 'Norm 30s Segments', 'Norm 30s Ave',
            'Norm 30s Coro', 'MORE Segmental changeOS', 'MORE changeOS',
            'MORE Coro changeOS', 'HV MORE Segmental changeOS',
            'HV MORE changeOS', 'HV MORE Coro changeOS']

  for patient in patients:
    reset=True
    id_to_check=1
    rows_ex.append([patient])
    biomarkers_patient=list(f_export[patient].keys())
    for marker in biomarkers_patient:
      if not reset or marker not in req_markers:
        break
      for file in files:
        if file not in f_export[patient][marker]:
          break
        if not reset:
          break
        for value_id in f_export[patient][marker][file]:
          if not reset:
            break
          if 'Segment' in value_id:
            for _,val in enumerate(f_export[patient][marker][file][value_id]):
              if not reset:
                break
              rows_ex[-1].append(val)

              if first_patient:
                rows_ex[0].append(marker.replace('Biomarker: ',''))
                rows_ex[1].append(file+' '+value_id)
                rows_ex[2].append('AHA'+str((_+1)))
              elif rows_ex[1][id_to_check]!=(file+' '+value_id):
                del rows_ex[-1]
                reset= False
                break

              else:
                id_to_check+=1

          else:
            for val_i in f_export[patient][marker][file][value_id]:
              rows_ex[-1].append(f_export[patient][marker][file][value_id][val_i])
              if first_patient:
                rows_ex[0].append(marker.replace('Biomarker: ',''))
                rows_ex[1].append(file+' '+value_id)
                rows_ex[2].append(val_i)
              elif rows_ex[1][id_to_check]!=(file+' '+value_id):
                del rows_ex[-1]
                reset= False
                break
                #print(patient)
                #1/0

              else:
                id_to_check+=1
    first_patient=False
  for row in rows_ex:
    ws_out_1.append(row)
        
  wb_output_1.save(filename)

def run(markers,input_files,phase_file,filename,required_phase=[1,1,1],get_frame=0):
  
  phase_info=get_phase(BIOMARKERS_INPUT,phase_file,input_files,required_phase)
  all_ex=get_raw_all(input_files,phase_info,required_phase)
  f_export=ready_export_to_excel(all_ex)
  data_frame=get_data_frame(all_ex)
  export_to_excel(f_export,filename,markers)
  if get_frame:
    data_frame.to_csv(f'data_frame_{filename}.csv')
  
  return f_export,data_frame
#phase_info=get_phase(BIOMARKERS_INPUT,phase_file,input_files,[1,0,0])
#all_ex=get_raw_all(input_files,phase_info,[1,0,0])
#f_export=ready_export_to_excel(all_ex)
#data_frame=get_data_frame(all_ex)
#import pickle
#with open('data_frame_export.pickle', 'wb') as handle:
    #pickle.dump(data_frame, handle)
#with open('data_export.pickle', 'wb') as handle:
    #pickle.dump(f_export, handle)

#export_to_excel(f_export,'filename',markers)
